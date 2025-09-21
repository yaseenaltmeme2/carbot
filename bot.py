# bot.py — مقفول على مجموعة واحدة + حذف تلقائي بعد 5 دقائق
import os, logging, asyncio, traceback
from typing import List, Dict, Optional
from openpyxl import load_workbook
from telegram import Update, Message
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from telegram.error import BadRequest, Forbidden, TimedOut, NetworkError, RetryAfter

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)

# ===== إعداد القفل والزمن =====
GROUP_ID = -1001234567890      # << بدّلها برقم مجموعتك بعد ما تستخدم /id داخل المجموعة
AUTO_DELETE_SECONDS = 300       # مدة الحذف التلقائي بالثواني (5 دقائق)
DELETE_BOT_MESSAGES = True      # إذا True يحذف حتى ردود البوت بعد المدة

def in_allowed_chat(update: Update) -> bool:
    return bool(update.effective_chat and update.effective_chat.id == GROUP_ID)

# ===== مساعدة: جدولة حذف رسالة =====
async def _delete_message(context: ContextTypes.DEFAULT_TYPE):
    chat_id, msg_id = context.job.data
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
    except Exception:
        pass

def schedule_autodelete(context: ContextTypes.DEFAULT_TYPE, msg: Optional[Message]):
    if not msg: return
    try:
        # ما نحذف إذا مو بالمجموعة المقفولة
        if msg.chat_id != GROUP_ID:
            return
        context.job_queue.run_once(_delete_message, AUTO_DELETE_SECONDS, data=(msg.chat_id, msg.message_id))
    except Exception:
        pass

# ===== ملفات البيانات (إكسل) =====
BASE = os.path.dirname(os.path.abspath(__file__))

# إذا تستخدم Persistent Disk على Render خلي DATA_DIR=/data بالـ Environment
DATA_DIR = os.getenv("DATA_DIR", BASE)

EXCEL_FILES = [
    os.path.join(DATA_DIR, "اكسل ارشيف باجات السيارات 2024.xlsx"),
    os.path.join(DATA_DIR, "اكسل سيارات 2025.xlsx"),
]

PLATE_CANDIDATES = ["رقمها", "رقم اللوحة", "رقم السيارة", "رقم العجلة", "رقم", "الرقم", "لوحة"]
RESPONSE_COLUMNS = ["الاسم الثلاثي", "العنوان", "الوظيفي", "مكان العمل", "لونها", "رقمها", "تسلسل الباج", "رقم الهاتف"]
MAX_LEN = 3900

def read_token():
    p = os.path.join(BASE, "token.txt")
    if os.path.exists(p):
        return open(p, "r", encoding="utf-8").read().strip()
    return os.getenv("TELEGRAM_BOT_TOKEN")

def norm(s: str) -> str:
    return str(s).replace("\u200f","").replace("\u200e","").strip().upper()

def norm_col(s: str) -> str:
    return str(s).replace("\u200f","").replace("\u200e","").strip()

def detect_plate_col(headers: List[str]) -> Optional[str]:
    headers_n = [norm_col(h) for h in headers]
    for k in PLATE_CANDIDATES[:4]:
        if k in headers_n:
            return headers[headers_n.index(k)]
    for i, h in enumerate(headers_n):
        for k in PLATE_CANDIDATES:
            if k in h:
                return headers[i]
    return None

def read_first_sheet_headers(ws) -> List[str]:
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [norm_col(c if c is not None else "") for c in row]
    return []

def build_row_dict(headers: List[str], row_values: List[str]) -> Dict[str, str]:
    out = {}
    for h, v in zip(headers, row_values):
        out[h] = "" if v is None else str(v)
    return out

def format_response(row: Dict[str,str], source_name:str) -> str:
    parts = []
    if "رقمها" in row:
        parts.append(f"🔎 نتيجة البحث للرقم: {row.get('رقمها','')}")
        parts.append("—" * 10)
    for col in RESPONSE_COLUMNS:
        if col in row:
            parts.append(f"{col}: {row.get(col,'')}")
    parts.append(f"المصدر: {source_name}")
    return "\n".join(parts)

def search_plate_once(xlsx_path: str, key: str) -> Optional[Dict[str,str]]:
    if not os.path.exists(xlsx_path):
        return None
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = read_first_sheet_headers(ws)
        if not headers:
            wb.close(); return None
        plate_col_name = detect_plate_col(headers)
        if not plate_col_name:
            wb.close(); return None
        headers_map = {headers[i]: i for i in range(len(headers))}
        plate_idx = headers_map.get(plate_col_name, None)
        if plate_idx is None:
            wb.close(); return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or plate_idx >= len(row): continue
            val = row[plate_idx]
            if val is None: continue
            if norm(str(val)) == key or (key in norm(str(val))):
                row_dict = build_row_dict(headers, list(row))
                wb.close()
                return row_dict
        wb.close()
        return None
    except Exception as e:
        logging.exception(f"خطأ أثناء قراءة {xlsx_path}: {e}")
        return None

# ===== إرسال آمن + حذف ردود البوت تلقائيًا =====
async def safe_send_text(update: Update, text: str, context: ContextTypes.DEFAULT_TYPE, max_attempts=3):
    attempt = 0
    msg = None
    while attempt < max_attempts:
        try:
            msg = await update.message.reply_text(text)
            break
        except RetryAfter as e:
            await asyncio.sleep(e.retry_after + 1); attempt += 1
        except (BadRequest, Forbidden, TimedOut, NetworkError):
            await asyncio.sleep(1.2); attempt += 1
    if msg and DELETE_BOT_MESSAGES:
        schedule_autodelete(context, msg)
    return msg

async def send_in_chunks(update: Update, text: str, context: ContextTypes.DEFAULT_TYPE):
    start, n = 0, len(text)
    while start < n:
        end = min(start + MAX_LEN, n)
        msg = await safe_send_text(update, text[start:end], context)
        start = end

# ===== أوامر (مقفولة بالتسجيل على المجموعة فقط) =====
async def id_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # مسموح فقط داخل المجموعة
    await safe_send_text(update,
        f"User ID: {update.effective_user.id}\nChat ID: {update.effective_chat.id}",
        context
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await safe_send_text(update,
        "👋 أهلاً بأعضاء المجموعة.\n"
        "أرسل رقم السيارة للبحث في السجلات.\n"
        "أوامر: /ping /id",
        context
    )

async def ping(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await safe_send_text(update, "pong ✅", context)

async def debug_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    existing = [os.path.basename(f) for f in EXCEL_FILES if os.path.exists(f)]
    msg = ["Files on server:"] + existing if existing else ["ماكو ملفات إكسل على السيرفر."]
    await safe_send_text(update, "\n".join(msg), context)

# ===== جامع الرسائل: يسجّل الحذف التلقائي لكل رسالة بالمجموعة =====
async def collect_and_autodelete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    أي رسالة داخل المجموعة: نسجّلها للحذف بعد AUTO_DELETE_SECONDS.
    - نحذف رسائل الأعضاء.
    - إذا رسالة من بوت ثاني/سيستم قد تفشل—نتجاهل الخطأ.
    """
    msg = update.effective_message
    if not msg: return
    # لا تحذف رسائل البوت مباشرةً هنا (نحذفها عبر safe_send_text حسب الإعداد)
    if msg.from_user and msg.from_user.is_bot:
        return
    schedule_autodelete(context, msg)

# ===== البحث =====
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = (update.message.text or "").strip()
        if not text:
            await safe_send_text(update, "اكتب رقم السيارة حتى أبحث عنه.", context)
            return
        key = norm(text)

        for path in EXCEL_FILES:
            row = search_plate_once(path, key)
            if row:
                msg = format_response(row, source_name=os.path.basename(path))
                if len(msg) > MAX_LEN:
                    await send_in_chunks(update, msg, context)
                else:
                    await safe_send_text(update, msg, context)
                return

        await safe_send_text(update, f"ماكو معلومات للسيارة رقم: {text}", context)

    except (BadRequest, Forbidden, TimedOut, NetworkError, RetryAfter) as e:
        logging.error(f"Telegram error: {type(e).__name__}: {e}")
        await safe_send_text(update, "خطأ إرسال: جرّب بعد شوي.", context)
    except Exception:
        logging.error("Unhandled error:\n" + traceback.format_exc())
        await safe_send_text(update, "صار خطأ غير متوقع داخل البوت.", context)

# ===== التشغيل =====
if __name__ == "__main__":
    token = read_token()
    if not token:
        print("ضع التوكن في TELEGRAM_BOT_TOKEN أو token.txt.")
        raise SystemExit(1)

    app = ApplicationBuilder().token(token).build()

    # ✅ تسجيـل الأوامر والنصوص *محصورة على نفس المجموعة فقط*:
    only_group = filters.Chat(GROUP_ID)

    app.add_handler(CommandHandler("start", start, filters=only_group))
    app.add_handler(CommandHandler("ping", ping, filters=only_group))
    app.add_handler(CommandHandler("id", id_cmd, filters=only_group))
    app.add_handler(CommandHandler("debug", debug_cmd, filters=only_group))

    # يجمع كل الرسائل داخل المجموعة لجدولة حذفها (قبل أي هاندلر ثانية)
    app.add_handler(MessageHandler(only_group & filters.ALL, collect_and_autodelete))

    # نصوص البحث داخل المجموعة فقط
    app.add_handler(MessageHandler(only_group & (filters.TEXT & ~filters.COMMAND), handle_message))

    print("البوت يعمل داخل المجموعة فقط. سيتم حذف الرسائل تلقائيًا بعد المدة المحددة.")
    app.run_polling()
